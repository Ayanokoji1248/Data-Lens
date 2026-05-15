import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any


NULL_STRINGS = {"", "null", "none", "nan", "n/a", "na"}
TRUE_STRINGS = {"true", "yes", "y", "1"}
FALSE_STRINGS = {"false", "no", "n", "0"}
DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y")
DATETIME_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S")
TYPE_ORDER = ("boolean", "integer", "float", "date", "datetime", "text")


@dataclass
class ParsedSpreadsheet:
    columns: list[dict[str, Any]]
    rows: list[dict[str, Any]]
    row_count: int
    column_count: int
    sheet_count: int
    sheet_names: list[str]


def parse_spreadsheet(path: Path, extension: str, sample_size: int) -> ParsedSpreadsheet:
    if extension == ".csv":
        headers, raw_rows = _read_csv(path)
        return _build_parsed_spreadsheet(headers, raw_rows, sample_size, sheet_names=[])

    if extension == ".xlsx":
        headers, raw_rows, sheet_names = _read_xlsx(path)
        return _build_parsed_spreadsheet(headers, raw_rows, sample_size, sheet_names=sheet_names)

    raise ValueError("Legacy .xls parsing is not supported yet.")


def _read_csv(path: Path) -> tuple[list[Any], list[list[Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file)
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise ValueError("Uploaded CSV is empty.") from exc
        return headers, [row for row in reader]


def _read_xlsx(path: Path) -> tuple[list[Any], list[list[Any]], list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        raise ValueError("Uploaded workbook does not contain any sheets.")

    worksheet = workbook[sheet_names[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("First worksheet is empty.")

    headers = list(rows[0])
    data_rows = [list(row) for row in rows[1:]]
    workbook.close()
    return headers, data_rows, sheet_names


def _build_parsed_spreadsheet(
    headers: list[Any],
    raw_rows: list[list[Any]],
    sample_size: int,
    sheet_names: list[str],
) -> ParsedSpreadsheet:
    if not headers or all(_is_null(header) for header in headers):
        raise ValueError("Uploaded file must include a header row.")

    columns = _build_columns(headers)
    normalized_rows = [_normalize_row(row, len(columns)) for row in raw_rows]

    for index, column in enumerate(columns):
        values = [row[index] for row in normalized_rows]
        inferred_type = _infer_type(values)
        column["inferredType"] = inferred_type
        column["nullCount"] = sum(1 for value in values if _is_null(value))
        column["sampleValues"] = _sample_values(values, sample_size)

    rows = []
    for raw_row in normalized_rows:
        rows.append(
            {
                column["storedName"]: _coerce_value(raw_row[index], column["inferredType"])
                for index, column in enumerate(columns)
            }
        )

    return ParsedSpreadsheet(
        columns=columns,
        rows=rows,
        row_count=len(rows),
        column_count=len(columns),
        sheet_count=len(sheet_names) if sheet_names else 1,
        sheet_names=sheet_names,
    )


def _build_columns(headers: list[Any]) -> list[dict[str, Any]]:
    used_names: dict[str, int] = {}
    columns = []

    for index, header in enumerate(headers):
        original_name = str(header).strip() if not _is_null(header) else f"Column {index + 1}"
        base_name = _sanitize_column_name(original_name) or f"column_{index + 1}"
        count = used_names.get(base_name, 0)
        used_names[base_name] = count + 1
        stored_name = base_name if count == 0 else f"{base_name}_{count + 1}"

        columns.append(
            {
                "originalName": original_name,
                "storedName": stored_name,
                "position": index,
            }
        )

    return columns


def _normalize_row(row: list[Any], column_count: int) -> list[Any]:
    normalized = list(row[:column_count])
    if len(normalized) < column_count:
        normalized.extend([None] * (column_count - len(normalized)))
    return normalized


def _sanitize_column_name(value: str) -> str:
    sanitized = re.sub(r"[^0-9a-zA-Z_]+", "_", value.strip().lower()).strip("_")
    if sanitized and sanitized[0].isdigit():
        sanitized = f"col_{sanitized}"
    return sanitized


def _infer_type(values: list[Any]) -> str:
    non_null_values = [value for value in values if not _is_null(value)]
    if not non_null_values:
        return "text"

    for type_name in TYPE_ORDER[:-1]:
        if all(_can_parse(value, type_name) for value in non_null_values):
            return type_name

    return "text"


def _can_parse(value: Any, type_name: str) -> bool:
    try:
        return _coerce_value(value, type_name) is not None
    except (TypeError, ValueError):
        return False


def _coerce_value(value: Any, type_name: str) -> Any:
    if _is_null(value):
        return None

    if isinstance(value, datetime):
        if type_name == "date":
            return value.date()
        if type_name == "datetime":
            return value
    if isinstance(value, date) and type_name in {"date", "datetime"}:
        return value

    text_value = str(value).strip()

    if type_name == "boolean":
        lowered = text_value.lower()
        if lowered in TRUE_STRINGS:
            return True
        if lowered in FALSE_STRINGS:
            return False
        raise ValueError("Invalid boolean value.")

    if type_name == "integer":
        return int(text_value)

    if type_name == "float":
        return float(text_value)

    if type_name == "date":
        return _parse_date(text_value)

    if type_name == "datetime":
        return _parse_datetime(text_value)

    return text_value


def _parse_date(value: str) -> date:
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return date.fromisoformat(value)


def _parse_datetime(value: str) -> datetime:
    for datetime_format in DATETIME_FORMATS:
        try:
            return datetime.strptime(value, datetime_format)
        except ValueError:
            continue
    return datetime.fromisoformat(value)


def _sample_values(values: list[Any], sample_size: int) -> list[Any]:
    samples = []
    for value in values:
        if _is_null(value):
            continue
        sample = value.isoformat() if isinstance(value, (date, datetime)) else value
        if sample not in samples:
            samples.append(sample)
        if len(samples) >= sample_size:
            break
    return samples


def _is_null(value: Any) -> bool:
    return value is None or str(value).strip().lower() in NULL_STRINGS
